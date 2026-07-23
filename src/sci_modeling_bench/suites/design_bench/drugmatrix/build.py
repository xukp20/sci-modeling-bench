"""Build the canonical DrugMatrix clinical-pathology release."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
from collections import defaultdict
from collections.abc import Iterable
from importlib import resources
from pathlib import Path
from typing import Any

from datasets import Dataset as HFDataset
from datasets import Features, Value
from huggingface_hub import DatasetCard
from huggingface_hub.repocard_data import DatasetCardData

from sci_modeling_bench.suites.design_bench.drugmatrix._conditions import (
    ENDPOINTS,
    build_measured_pool,
)
from sci_modeling_bench.suites.design_bench.drugmatrix.dataset import (
    DRUGMATRIX_CONFIG_NAME,
    DRUGMATRIX_DEFAULT_SPLIT,
    DrugMatrixValidator,
)

SOURCE_PAGE = "https://cebs.niehs.nih.gov/cebs/paper/15670"
SOURCE_DOI = "10.22427/NTP-DATA-107-022-001-000-3"
CLINICAL_PATHOLOGY_FILENAME = "DrugMatrix_ClinicalChemistry_Hematology.xlsx"
CLINICAL_PATHOLOGY_SIZE = 2_967_840
CLINICAL_PATHOLOGY_SHA256 = (
    "0cc3a2aabe8585cab34c920d9d46d7323f1529d1caff452143c2d1ac818fc0fd"
)
CURATED_CHEMICALS_FILENAME = "Drugmatrix_Curated_Chemicals_With_SMILES.txt"
CURATED_CHEMICALS_SIZE = 542_925
CURATED_CHEMICALS_SHA256 = (
    "fa9509d438290e37059e1bcfb75d96c6fa500ec8a9a22ca8829b299c05f133bd"
)
CHEMBL_ASSAY_ID = "CHEMBL3885882"
CHEMBL_ACTIVITY_SHA256 = (
    "c1349b3d65340e5d788db7f84c7807f7b9cef5a9a2de822a9dca27e927bd83f1"
)

DATASET_ID = "design-bench/drugmatrix-clinical-pathology"
DATASET_VERSION = "1.0.0"
SOURCE_SHEET = "DM_ClinicalChemistry_Hematology"
EXPECTED_SOURCE_ROWS = 10_605
EXPECTED_SOURCE_COLUMNS = 53

_SOURCE_TO_RELEASE = {
    "ANIMAL_ID": "animal_id",
    "STUDY_ID": "study_id",
    "CASRN": "casrn",
    "DOSE": "dose",
    "DURATION": "duration_days",
    "VEHICLE": "vehicle",
    "TREATMENT": "treatment",
    "SEX": "sex",
    "ROUTE": "route",
    "MCHC": "mchc",
    "MCH": "mch",
    "CREATININE": "creatinine",
    "NA": "sodium",
    "CHLORIDE": "chloride",
    "PHOSPHORUS": "phosphorus",
}
_NUMERIC_FIELDS = {
    "dose",
    "duration_days",
    *ENDPOINTS,
}
_NAME_PATTERN = re.compile(r"[^a-z0-9]+")

KNOWLEDGE_RESOURCES = {
    "smiles_molecular_graphs_and_chemical_identity": {
        "title": "SMILES, Molecular Graphs, and Chemical Identity",
        "description": (
            "SMILES atoms, bonds, branches, rings, aromaticity, charge, "
            "stereochemistry, disconnected components, and canonicalization."
        ),
        "source_path": "shared/smiles-molecular-graphs-and-chemical-identity.md",
        "path": "knowledge/shared/smiles-molecular-graphs-and-chemical-identity.md",
        "media_type": "text/markdown",
    },
    "repeated_dose_toxicology_and_experimental_controls": {
        "title": "Repeated-Dose Toxicology and Experimental Controls",
        "description": (
            "Repeated-dose rodent study design, treatment groups, concurrent "
            "controls, biological replication, clinical pathology, and interpretation."
        ),
        "source_path": (
            "shared/repeated-dose-toxicology-and-experimental-controls.md"
        ),
        "path": (
            "knowledge/shared/"
            "repeated-dose-toxicology-and-experimental-controls.md"
        ),
        "media_type": "text/markdown",
    },
    "toxicokinetics_dose_route_duration_and_vehicle": {
        "title": "Toxicokinetics: Dose, Route, Duration, and Vehicle",
        "description": (
            "Administered dose versus systemic exposure, ADME, dose response, "
            "route, repeated exposure, sex, and formulation or vehicle effects."
        ),
        "source_path": (
            "shared/toxicokinetics-dose-route-duration-and-vehicle.md"
        ),
        "path": (
            "knowledge/shared/"
            "toxicokinetics-dose-route-duration-and-vehicle.md"
        ),
        "media_type": "text/markdown",
    },
    "laboratory_rat_clinical_pathology": {
        "title": "Laboratory Rat Clinical Pathology",
        "description": (
            "Clinical pathology measurements in rats, biological and analytical "
            "variation, matched controls, reference intervals, and adversity."
        ),
        "source_path": "shared/laboratory-rat-clinical-pathology.md",
        "path": "knowledge/shared/laboratory-rat-clinical-pathology.md",
        "media_type": "text/markdown",
    },
    "erythrocyte_indices_mch_and_mchc": {
        "title": "Erythrocyte Indices: MCH and MCHC",
        "description": (
            "Definitions, formulas, units, dependencies, biological interpretation, "
            "and analytical limitations of MCH and MCHC."
        ),
        "source_path": "shared/erythrocyte-indices-mch-and-mchc.md",
        "path": "knowledge/shared/erythrocyte-indices-mch-and-mchc.md",
        "media_type": "text/markdown",
    },
    "creatinine_and_renal_function": {
        "title": "Creatinine and Renal Function",
        "description": (
            "Creatinine production and excretion, relation to glomerular filtration, "
            "nonrenal determinants, sensitivity, and toxicologic interpretation."
        ),
        "source_path": "shared/creatinine-and-renal-function.md",
        "path": "knowledge/shared/creatinine-and-renal-function.md",
        "media_type": "text/markdown",
    },
    "sodium_chloride_and_fluid_acid_base_homeostasis": {
        "title": "Sodium, Chloride, and Fluid and Acid-Base Homeostasis",
        "description": (
            "Extracellular sodium and chloride, tonicity, water balance, renal and "
            "hormonal regulation, acid-base coupling, and measurement interpretation."
        ),
        "source_path": (
            "shared/sodium-chloride-and-fluid-acid-base-homeostasis.md"
        ),
        "path": (
            "knowledge/shared/"
            "sodium-chloride-and-fluid-acid-base-homeostasis.md"
        ),
        "media_type": "text/markdown",
    },
    "phosphorus_homeostasis": {
        "title": "Phosphorus Homeostasis",
        "description": (
            "Circulating inorganic phosphate, intestinal absorption, bone exchange, "
            "renal handling, and regulation by PTH, FGF23, and vitamin D."
        ),
        "source_path": "shared/phosphorus-homeostasis.md",
        "path": "knowledge/shared/phosphorus-homeostasis.md",
        "media_type": "text/markdown",
    },
}


def build_drugmatrix_release(
    clinical_pathology: str | Path,
    curated_chemicals: str | Path,
    chembl_activity: str | Path,
    output_dir: str | Path,
) -> dict[str, Any]:
    """Build one upload-ready DrugMatrix config and return its provenance."""

    clinical_path = Path(clinical_pathology)
    curated_path = Path(curated_chemicals)
    chembl_path = Path(chembl_activity)
    destination = Path(output_dir)
    _verify_source(
        clinical_path,
        expected_name=CLINICAL_PATHOLOGY_FILENAME,
        expected_size=CLINICAL_PATHOLOGY_SIZE,
        expected_sha256=CLINICAL_PATHOLOGY_SHA256,
    )
    _verify_source(
        curated_path,
        expected_name=CURATED_CHEMICALS_FILENAME,
        expected_size=CURATED_CHEMICALS_SIZE,
        expected_sha256=CURATED_CHEMICALS_SHA256,
    )
    _verify_source(
        chembl_path,
        expected_name=chembl_path.name,
        expected_size=None,
        expected_sha256=CHEMBL_ACTIVITY_SHA256,
    )

    crosswalk, crosswalk_report = build_structure_crosswalk(
        curated_path,
        chembl_path,
    )
    rows, source_statistics = canonicalize_clinical_pathology(
        clinical_path,
        crosswalk,
    )
    treatment_articles = source_statistics.pop("_treatment_articles")
    crosswalk_report = _restrict_crosswalk_report(
        crosswalk_report,
        treatment_articles,
    )
    data = HFDataset.from_dict(rows, features=_features())
    validation = DrugMatrixValidator().validate_dataset(data)
    if not validation.valid:
        raise ValueError(
            "canonical DrugMatrix validation failed: "
            + "; ".join(item.message for item in validation.violations[:10])
        )

    data_path = (
        destination
        / "data"
        / DRUGMATRIX_CONFIG_NAME
        / f"{DRUGMATRIX_DEFAULT_SPLIT}.parquet"
    )
    data_path.parent.mkdir(parents=True, exist_ok=True)
    data.to_parquet(data_path)

    pool = build_measured_pool(data).table
    pool_statistics = _pool_statistics(pool)
    provenance = _provenance(
        clinical_path,
        curated_path,
        chembl_path,
        data_path,
        source_statistics,
        crosswalk_report,
        pool_statistics,
    )
    provenance["knowledge"] = _write_knowledge_resources(destination)
    _write_release_metadata(destination, provenance)
    return provenance


def canonicalize_clinical_pathology(
    source: str | Path,
    crosswalk: dict[str, dict[str, Any]],
) -> tuple[dict[str, list[Any]], dict[str, Any]]:
    """Select canonical columns while preserving individual-animal rows."""

    from openpyxl import load_workbook

    workbook = load_workbook(source, read_only=True, data_only=True)
    if SOURCE_SHEET not in workbook.sheetnames:
        raise ValueError(f"source workbook is missing sheet {SOURCE_SHEET!r}")
    worksheet = workbook[SOURCE_SHEET]
    iterator = worksheet.iter_rows(values_only=True)
    header = tuple(next(iterator))
    if len(header) != EXPECTED_SOURCE_COLUMNS:
        raise ValueError(
            f"expected {EXPECTED_SOURCE_COLUMNS} source columns, got {len(header)}"
        )
    missing = sorted(set(_SOURCE_TO_RELEASE) - set(header))
    if missing:
        raise ValueError(f"source workbook is missing required columns: {missing}")
    indices = {name: header.index(name) for name in _SOURCE_TO_RELEASE}
    test_article_index = header.index("TEST_ARTICLE")
    species_index = header.index("SPECIES")
    strain_index = header.index("STRAIN")

    columns = {
        "animal_id": [],
        "study_id": [],
        "casrn": [],
        "canonical_smiles": [],
        "dose": [],
        "duration_days": [],
        "vehicle": [],
        "treatment": [],
        "sex": [],
        "route": [],
        **{endpoint: [] for endpoint in ENDPOINTS},
    }
    treatment_articles: dict[str, dict[str, set[str]]] = {}
    mapped_articles: set[str] = set()
    for source_row, row in enumerate(iterator, start=2):
        record: dict[str, Any] = {}
        for source_name, release_name in _SOURCE_TO_RELEASE.items():
            value = _clean_cell(row[indices[source_name]])
            if release_name in _NUMERIC_FIELDS:
                value = _number(value, release_name, source_row)
            record[release_name] = value

        if row[species_index] != "Rat" or row[strain_index] != "Sprague Dawley":
            raise ValueError(f"unexpected species/strain at source row {source_row}")
        if record["treatment"] == "Chemical":
            article = _clean_cell(row[test_article_index])
            if not isinstance(article, str) or not article:
                raise ValueError(f"chemical row {source_row} has no TEST_ARTICLE")
            normalized = normalize_name(article)
            article_record = treatment_articles.setdefault(
                normalized,
                {"source_test_articles": set(), "source_casrn": set()},
            )
            article_record["source_test_articles"].add(article)
            casrn = record["casrn"]
            if isinstance(casrn, str) and casrn:
                article_record["source_casrn"].add(casrn)
            mapping = crosswalk.get(normalized)
            record["canonical_smiles"] = (
                mapping["canonical_smiles"] if mapping is not None else None
            )
            if mapping is not None:
                mapped_articles.add(normalized)
        else:
            record["canonical_smiles"] = None

        for name in columns:
            columns[name].append(record[name])

    if len(columns["animal_id"]) != EXPECTED_SOURCE_ROWS:
        raise ValueError(
            f"expected {EXPECTED_SOURCE_ROWS} source rows, "
            f"got {len(columns['animal_id'])}"
        )
    complete = sum(
        all(columns[endpoint][index] is not None for endpoint in ENDPOINTS)
        for index in range(EXPECTED_SOURCE_ROWS)
    )
    return columns, {
        "rows": EXPECTED_SOURCE_ROWS,
        "unique_animal_ids": len(set(columns["animal_id"])),
        "chemical_test_articles": len(treatment_articles),
        "mapped_test_articles": len(mapped_articles),
        "unmapped_test_articles": len(set(treatment_articles) - mapped_articles),
        "rows_with_all_six_endpoints": complete,
        "_treatment_articles": treatment_articles,
    }


def build_structure_crosswalk(
    curated_chemicals: str | Path,
    chembl_activity: str | Path,
) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    """Build a deterministic exact-name structure crosswalk without fuzzy joins."""

    curated = _curated_structures(Path(curated_chemicals))
    chembl = _chembl_structures(Path(chembl_activity))
    names = sorted(set(curated) | set(chembl))
    selected: dict[str, dict[str, Any]] = {}
    entries: list[dict[str, Any]] = []
    method_counts: dict[str, int] = defaultdict(int)
    for name in names:
        chembl_values = chembl.get(name, set())
        curated_values = curated.get(name, set())
        method: str | None = None
        value: str | None = None
        if len(chembl_values) == 1:
            method = "chembl_exact_name"
            value = next(iter(chembl_values))
        elif not chembl_values and len(curated_values) == 1:
            method = "cebs_curated_exact_name"
            value = next(iter(curated_values))
        elif len(curated_values) == 1:
            method = "cebs_curated_exact_name_after_ambiguous_chembl"
            value = next(iter(curated_values))

        status = "mapped" if value is not None else "ambiguous"
        if value is not None and method is not None:
            selected[name] = {
                "canonical_smiles": value,
                "mapping_method": method,
            }
            method_counts[method] += 1
        entries.append(
            {
                "normalized_name": name,
                "canonical_smiles": value,
                "mapping_method": method,
                "status": status,
                "chembl_candidates": sorted(chembl_values),
                "cebs_curated_candidates": sorted(curated_values),
            }
        )
    return selected, {
        "normalization": "casefold and remove non-ASCII-alphanumeric characters",
        "fuzzy_matching": False,
        "selected_names": len(selected),
        "method_counts": dict(sorted(method_counts.items())),
        "entries": entries,
    }


def normalize_name(value: str) -> str:
    """Return the frozen exact-name join key."""

    return _NAME_PATTERN.sub("", value.casefold())


def _restrict_crosswalk_report(
    report: dict[str, Any],
    treatment_articles: dict[str, dict[str, set[str]]],
) -> dict[str, Any]:
    by_name = {entry["normalized_name"]: entry for entry in report["entries"]}
    entries: list[dict[str, Any]] = []
    method_counts: dict[str, int] = defaultdict(int)
    for name in sorted(treatment_articles):
        entry = dict(
            by_name.get(
                name,
                {
                    "normalized_name": name,
                    "canonical_smiles": None,
                    "mapping_method": None,
                    "status": "unmapped",
                    "chembl_candidates": [],
                    "cebs_curated_candidates": [],
                },
            )
        )
        source = treatment_articles[name]
        entry["source_test_articles"] = sorted(source["source_test_articles"])
        entry["source_casrn"] = sorted(source["source_casrn"])
        if entry["mapping_method"] is not None:
            method_counts[entry["mapping_method"]] += 1
        entries.append(entry)
    return {
        "normalization": report["normalization"],
        "fuzzy_matching": False,
        "selected_names": sum(entry["status"] == "mapped" for entry in entries),
        "method_counts": dict(sorted(method_counts.items())),
        "entries": entries,
    }


def _curated_structures(path: Path) -> dict[str, set[str]]:
    structures: dict[str, set[str]] = defaultdict(set)
    with path.open(encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file, delimiter="\t")
        if reader.fieldnames != ["COMPOUND", "COMPOUND_NAME", "SMILES"]:
            raise ValueError(f"unexpected curated-chemical columns: {reader.fieldnames}")
        for row in reader:
            name = row["COMPOUND_NAME"].strip()
            try:
                smiles = _canonical_smiles(
                    row["SMILES"], source="CEBS curated chemicals"
                )
            except ValueError:
                continue
            structures[normalize_name(name)].add(smiles)
    return structures


def _chembl_structures(path: Path) -> dict[str, set[str]]:
    activities = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(activities, list):
        raise ValueError("ChEMBL activity export must contain a JSON list")
    structures: dict[str, set[str]] = defaultdict(set)
    for activity in activities:
        name = activity.get("molecule_pref_name")
        smiles = activity.get("canonical_smiles")
        if not isinstance(name, str) or not name or not isinstance(smiles, str):
            continue
        try:
            canonical = _canonical_smiles(smiles, source="ChEMBL activity export")
        except ValueError:
            continue
        structures[normalize_name(name)].add(canonical)
    return structures


def _canonical_smiles(value: str, *, source: str) -> str:
    from rdkit import Chem

    molecule = Chem.MolFromSmiles(value)
    if molecule is None:
        raise ValueError(f"{source} contains an invalid SMILES: {value!r}")
    return str(Chem.MolToSmiles(molecule, canonical=True, isomericSmiles=True))


def _features() -> Features:
    return Features(
        {
            "animal_id": Value("string"),
            "study_id": Value("string"),
            "casrn": Value("string"),
            "canonical_smiles": Value("string"),
            "dose": Value("float64"),
            "duration_days": Value("float64"),
            "vehicle": Value("string"),
            "treatment": Value("string"),
            "sex": Value("string"),
            "route": Value("string"),
            **{endpoint: Value("float64") for endpoint in ENDPOINTS},
        }
    )


def _pool_statistics(pool: HFDataset) -> dict[str, Any]:
    result: dict[str, Any] = {
        "conditions": len(pool),
        "unique_casrn": len(set(pool["casrn"])),
        "unique_canonical_smiles": len(set(pool["canonical_smiles"])),
        "condition_id_sha256": _hash_values(pool["condition_id"]),
        "endpoints": {},
    }
    for endpoint in ENDPOINTS:
        raw = [float(value) for value in pool[f"{endpoint}_raw_response"]]
        scores = [float(value) for value in pool[f"{endpoint}_control_deviation"]]
        result["endpoints"][endpoint] = {
            "raw_response_minimum": min(raw),
            "raw_response_maximum": max(raw),
            "control_deviation_minimum": min(scores),
            "control_deviation_maximum": max(scores),
            "control_deviation_sha256": _hash_values(scores),
        }
    return result


def _provenance(
    clinical_pathology: Path,
    curated_chemicals: Path,
    chembl_activity: Path,
    data_path: Path,
    source_statistics: dict[str, Any],
    crosswalk: dict[str, Any],
    pool_statistics: dict[str, Any],
) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "dataset_id": DATASET_ID,
        "dataset_version": DATASET_VERSION,
        "config": DRUGMATRIX_CONFIG_NAME,
        "split": DRUGMATRIX_DEFAULT_SPLIT,
        "source_artifacts": [
            _artifact(clinical_pathology, SOURCE_PAGE),
            _artifact(curated_chemicals, SOURCE_PAGE),
            _artifact(
                chembl_activity,
                f"https://www.ebi.ac.uk/chembl/api/data/activity.json?assay_chembl_id={CHEMBL_ASSAY_ID}",
            ),
        ],
        "transformation": {
            "steps": [
                "verify all pinned source artifacts",
                "preserve the official individual-animal source row order",
                "select and rename the declared clinical-pathology fields",
                "canonicalize exact-name matched SMILES with RDKit",
                "retain unresolved chemical structures as null",
                "retain missing and extreme endpoint measurements without imputation",
            ],
            "structure_crosswalk": crosswalk,
        },
        "statistics": source_statistics,
        "measured_pool": pool_statistics,
        "artifact": {
            "path": (
                f"data/{DRUGMATRIX_CONFIG_NAME}/"
                f"{DRUGMATRIX_DEFAULT_SPLIT}.parquet"
            ),
            "size_bytes": data_path.stat().st_size,
            "sha256": _sha256(data_path),
        },
    }


def _write_release_metadata(destination: Path, provenance: dict[str, Any]) -> None:
    manifest_dir = destination / "manifests"
    provenance_dir = destination / "provenance" / DRUGMATRIX_CONFIG_NAME
    manifest_dir.mkdir(parents=True, exist_ok=True)
    provenance_dir.mkdir(parents=True, exist_ok=True)

    collection_path = destination / "scimodelingbench.json"
    if collection_path.exists():
        collection = json.loads(collection_path.read_text(encoding="utf-8"))
    else:
        collection = {
            "schema_version": 1,
            "default_config": DRUGMATRIX_CONFIG_NAME,
            "configs": {},
        }
    collection["configs"][DRUGMATRIX_CONFIG_NAME] = {
        "manifest": f"manifests/{DRUGMATRIX_CONFIG_NAME}.json"
    }

    manifest = {
        "schema_version": 1,
        "dataset_id": DATASET_ID,
        "version": DATASET_VERSION,
        "default_split": DRUGMATRIX_DEFAULT_SPLIT,
        "description": (
            "Individual-animal DrugMatrix clinical-pathology observations with "
            "audited molecular structures and six measured endpoints."
        ),
        "license": "unknown",
        "source": [
            {
                "name": "NIEHS CEBS DrugMatrix release",
                "url": SOURCE_PAGE,
                "version": SOURCE_DOI,
                "checksum": f"sha256:{CLINICAL_PATHOLOGY_SHA256}",
                "notes": "No explicit artifact redistribution license was identified.",
            },
            {
                "name": "ChEMBL DrugMatrix assay activity export",
                "url": (
                    "https://www.ebi.ac.uk/chembl/explore/assay/"
                    f"{CHEMBL_ASSAY_ID}"
                ),
                "version": CHEMBL_ASSAY_ID,
                "checksum": f"sha256:{CHEMBL_ACTIVITY_SHA256}",
            },
        ],
        "citation": [
            {
                "text": (
                    "National Toxicology Program, DrugMatrix toxicogenomic "
                    "database, DOI 10.22427/NTP-DATA-107-022-001-000-3."
                ),
                "url": f"https://doi.org/{SOURCE_DOI}",
            }
        ],
        "inputs": [
            {
                "name": "canonical_smiles",
                "description": "Audited canonical molecular SMILES when available.",
                "required": False,
            }
        ],
        "targets": [
            {
                "name": "mchc",
                "description": "Mean corpuscular hemoglobin concentration.",
                "unit": "g/dL",
                "required": False,
                "constraints": [{"kind": "finite"}],
            },
            {
                "name": "mch",
                "description": "Mean corpuscular hemoglobin.",
                "unit": "pg",
                "required": False,
                "constraints": [{"kind": "finite"}],
            },
            {
                "name": "creatinine",
                "description": "Creatinine concentration.",
                "unit": "mg/dL",
                "required": False,
                "constraints": [{"kind": "finite"}],
            },
            {
                "name": "sodium",
                "description": "Sodium concentration.",
                "unit": "mEq/L",
                "required": False,
                "constraints": [{"kind": "finite"}],
            },
            {
                "name": "chloride",
                "description": "Chloride concentration.",
                "unit": "mEq/L",
                "required": False,
                "constraints": [{"kind": "finite"}],
            },
            {
                "name": "phosphorus",
                "description": "Phosphorus concentration.",
                "unit": "mg/dL",
                "required": False,
                "constraints": [{"kind": "finite"}],
            },
        ],
        "context": [
            {
                "name": "animal_id",
                "description": "Unique source animal identifier.",
                "required": False,
            },
            {
                "name": "study_id",
                "description": "Source toxicology study identifier.",
                "required": False,
            },
            {
                "name": "casrn",
                "description": "CAS Registry Number for chemical treatments.",
                "required": False,
            },
            {
                "name": "dose",
                "description": "Administered source dose value.",
                "required": False,
                "constraints": [
                    {"kind": "finite"},
                    {"kind": "range", "minimum": 0.0},
                ],
            },
            {
                "name": "duration_days",
                "description": "Exposure duration in days.",
                "unit": "day",
                "required": False,
                "constraints": [
                    {"kind": "finite"},
                    {"kind": "range", "minimum": 0.0},
                ],
            },
            {
                "name": "vehicle",
                "description": "Administration vehicle.",
                "required": False,
            },
            {
                "name": "treatment",
                "description": "Chemical or control treatment class.",
                "required": False,
                "constraints": [
                    {
                        "kind": "allowed_values",
                        "values": [
                            "Chemical",
                            "Vehicle Control",
                            "Untreated Control",
                        ],
                    }
                ],
            },
            {
                "name": "sex",
                "description": "Animal sex when reported.",
                "required": False,
            },
            {
                "name": "route",
                "description": "Administration route when reported.",
                "required": False,
            },
        ],
        "splits": [
            {
                "name": DRUGMATRIX_DEFAULT_SPLIT,
                "description": "All canonical individual-animal observations.",
                "num_rows": EXPECTED_SOURCE_ROWS,
                "attributes": {
                    "species": "Rat",
                    "strain": "Sprague Dawley",
                    "source_granularity": "individual_animal",
                },
            }
        ],
        "knowledge": {
            key: {
                field: value
                for field, value in specification.items()
                if field != "source_path"
            }
            for key, specification in KNOWLEDGE_RESOURCES.items()
        },
    }
    _write_json(collection_path, collection)
    _write_json(manifest_dir / f"{DRUGMATRIX_CONFIG_NAME}.json", manifest)
    _write_json(
        provenance_dir / f"{DRUGMATRIX_DEFAULT_SPLIT}.json",
        provenance,
    )
    _update_dataset_card(destination / "README.md", provenance)


def _update_dataset_card(path: Path, provenance: dict[str, Any]) -> None:
    if path.exists():
        card = DatasetCard.load(path)
    else:
        card = DatasetCard(
            "---\nlicense: other\nconfigs: []\n---\n\n"
            "# SciModelingBench Design-Bench Data\n\n"
            "Canonical observations used by the Design-Bench suite.\n"
        )
    metadata = card.data.to_dict()
    configs = [
        config
        for config in metadata.get("configs", [])
        if config.get("config_name") != DRUGMATRIX_CONFIG_NAME
    ]
    configs.append(
        {
            "config_name": DRUGMATRIX_CONFIG_NAME,
            "data_files": [
                {
                    "split": DRUGMATRIX_DEFAULT_SPLIT,
                    "path": provenance["artifact"]["path"],
                }
            ],
        }
    )
    metadata["configs"] = configs
    metadata["license"] = "other"
    card.data = DatasetCardData(**metadata)

    marker = "## DrugMatrix Clinical Pathology"
    section = f"""
## DrugMatrix Clinical Pathology

The `{DRUGMATRIX_CONFIG_NAME}` config contains {EXPECTED_SOURCE_ROWS:,}
individual-animal rat clinical-pathology observations and six measured
endpoints. Molecular structures use an audited exact-name crosswalk; unresolved
treatments remain null. See
`provenance/{DRUGMATRIX_CONFIG_NAME}/{DRUGMATRIX_DEFAULT_SPLIT}.json`.

The source page does not state a simple redistribution license for these
artifacts, so the config manifest records `license: unknown`.
"""
    if marker not in card.text:
        card.text = card.text.rstrip() + "\n" + section
    card.save(path)


def _verify_source(
    path: Path,
    *,
    expected_name: str,
    expected_size: int | None,
    expected_sha256: str,
) -> None:
    if not path.is_file():
        raise FileNotFoundError(path)
    if path.name != expected_name:
        raise ValueError(f"expected source filename {expected_name!r}, got {path.name!r}")
    if expected_size is not None and path.stat().st_size != expected_size:
        raise ValueError(f"source size mismatch for {path.name}")
    if _sha256(path) != expected_sha256:
        raise ValueError(f"source checksum mismatch for {path.name}")


def _clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def _number(value: Any, field: str, source_row: int) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{field} at source row {source_row} must be numeric or null")
    result = float(value)
    if not math.isfinite(result):
        raise ValueError(f"{field} at source row {source_row} must be finite")
    return result


def _artifact(path: Path, url: str) -> dict[str, Any]:
    return {
        "filename": path.name,
        "url": url,
        "size_bytes": path.stat().st_size,
        "sha256": _sha256(path),
    }


def _hash_values(values: Iterable[Any]) -> str:
    encoded = json.dumps(
        list(values),
        ensure_ascii=True,
        allow_nan=False,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _write_knowledge_resources(destination: Path) -> list[dict[str, Any]]:
    source_root = resources.files("sci_modeling_bench").joinpath(
        "resources", "knowledge"
    )
    artifacts: list[dict[str, Any]] = []
    for key, specification in KNOWLEDGE_RESOURCES.items():
        content = source_root.joinpath(specification["source_path"]).read_bytes()
        output_path = destination / specification["path"]
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(content)
        artifacts.append(
            {
                "key": key,
                "path": specification["path"],
                "size_bytes": len(content),
                "sha256": hashlib.sha256(content).hexdigest(),
            }
        )
    return artifacts


def _write_json(path: Path, content: dict[str, Any]) -> None:
    path.write_text(
        json.dumps(content, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Build the DrugMatrix clinical-pathology HF config."
    )
    parser.add_argument("--clinical-pathology", type=Path, required=True)
    parser.add_argument("--curated-chemicals", type=Path, required=True)
    parser.add_argument("--chembl-activity", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    provenance = build_drugmatrix_release(
        args.clinical_pathology,
        args.curated_chemicals,
        args.chembl_activity,
        args.output_dir,
    )
    print(json.dumps(provenance, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
